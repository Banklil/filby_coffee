from fastapi import APIRouter, Depends, Query, Response
from sqlalchemy.orm import Session
from typing import Optional
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from ..database import get_db
from ..core.deps import get_current_user, require_roles
from ..models.admin import Admin
from ..models.order import Order
from ..models.transaction import Transaction
from ..models.shop import Shop

router = APIRouter(prefix="/api/reports", tags=["reports"])


def make_excel_response(wb: Workbook, filename: str) -> Response:
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(
        content=buf.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/orders.xlsx")
def export_orders(
    from_date: Optional[str] = Query(None, alias="from"),
    to_date: Optional[str] = Query(None, alias="to"),
    db: Session = Depends(get_db),
    current_user: Admin = Depends(require_roles("super_admin", "manager", "accountant")),
):
    q = db.query(Order)
    orders = q.order_by(Order.created_at.desc()).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Orders"
    headers = ["Order ID", "Shop ID", "Amount (LAK)", "Status", "Payment Method", "Created At"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="E8854A")

    for row, order in enumerate(orders, 2):
        ws.cell(row=row, column=1, value=order.order_id)
        ws.cell(row=row, column=2, value=order.shop_id)
        ws.cell(row=row, column=3, value=order.amount)
        ws.cell(row=row, column=4, value=order.status)
        ws.cell(row=row, column=5, value=order.payment_method)
        ws.cell(row=row, column=6, value=str(order.created_at))

    return make_excel_response(wb, "filby-orders.xlsx")


@router.get("/payments.xlsx")
def export_payments(
    db: Session = Depends(get_db),
    current_user: Admin = Depends(require_roles("super_admin", "manager", "accountant")),
):
    txns = db.query(Transaction).filter(Transaction.type == "payment").order_by(Transaction.created_at.desc()).all()
    wb = Workbook()
    ws = wb.active
    ws.title = "Payments"
    headers = ["ID", "Shop ID", "Amount (LAK)", "Description", "Date"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)

    for row, txn in enumerate(txns, 2):
        ws.cell(row=row, column=1, value=txn.id)
        ws.cell(row=row, column=2, value=txn.shop_id)
        ws.cell(row=row, column=3, value=txn.amount)
        ws.cell(row=row, column=4, value=txn.description)
        ws.cell(row=row, column=5, value=str(txn.created_at))

    return make_excel_response(wb, "filby-payments.xlsx")


@router.get("/shop/{shop_id}/statement")
def shop_statement(
    shop_id: int,
    month: str = Query(..., description="YYYY-MM"),
    db: Session = Depends(get_db),
    current_user: Admin = Depends(get_current_user),
):
    shop = db.query(Shop).filter(Shop.id == shop_id).first()
    if not shop:
        from fastapi import HTTPException
        raise HTTPException(status_code=404, detail="ບໍ່ພົບຮ້ານ")
    content = f"Statement for {shop.name}\nMonth: {month}\nCredit Used: {shop.credit_used:,} LAK"
    return Response(
        content=content.encode("utf-8"),
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=statement-{shop.shop_id}-{month}.pdf"},
    )
